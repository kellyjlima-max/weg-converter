import os
import io
import re
import json
import base64
from pathlib import Path

import anthropic
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, File, UploadFile, HTTPException, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# ─── App ─────────────────────────────────────────────────────────────────────
app = FastAPI(title="Conversor de Códigos WEG – THP")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

static_dir = Path("static")
static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

# ─── Claude ──────────────────────────────────────────────────────────────────
client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

VISION_MODEL = os.environ.get("VISION_MODEL", "claude-sonnet-4-5")
TEXT_MODEL   = os.environ.get("TEXT_MODEL",   "claude-sonnet-4-5")

# ─── Azure SQL — Conexão e Ferramenta ────────────────────────────────────────
import pymssql


def get_conn():
    return pymssql.connect(
        server=os.environ["AZURE_SQL_SERVER"],
        user=os.environ["AZURE_SQL_USER"],
        password=os.environ["AZURE_SQL_PASSWORD"],
        database=os.environ["AZURE_SQL_DB"],
        login_timeout=10,
        timeout=30,
        charset="UTF-8",
    )


def buscar_produto_weg(familia, corrente_min=None, corrente_max=None,
                       tensao_v=None, potencia_kvar=None, potencia_kw=None,
                       potencia_cv=None, texto_livre=None):
    """Consulta weg_produtos no Azure SQL e retorna JSON string."""
    try:
        conn = get_conn()
        cursor = conn.cursor(as_dict=True)

        clauses = ["ativo = 1", "familia = %s"]
        params = [familia]

        if corrente_min is not None:
            clauses.append("corrente_max >= %s")
            params.append(float(corrente_min))
        if corrente_max is not None:
            clauses.append("corrente_min <= %s")
            params.append(float(corrente_max))
        if tensao_v:
            clauses.append("tensao_v = %s")
            params.append(str(tensao_v))
        if potencia_kvar is not None:
            clauses.append("potencia_kvar >= %s")
            params.append(float(potencia_kvar) * 0.95)
        if potencia_kw is not None:
            clauses.append("potencia_kw >= %s")
            params.append(float(potencia_kw) * 0.95)
        if potencia_cv is not None:
            clauses.append("potencia_cv >= %s")
            params.append(float(potencia_cv) * 0.95)
        if texto_livre:
            clauses.append("(codigo LIKE %s OR subtipo LIKE %s OR observacoes LIKE %s)")
            like = "%" + texto_livre + "%"
            params.extend([like, like, like])

        sql = (
            "SELECT TOP 30 * FROM weg_produtos "
            "WHERE " + " AND ".join(clauses) + " "
            "ORDER BY corrente_min, potencia_kvar, potencia_kw, potencia_cv"
        )
        cursor.execute(sql, params)
        rows = cursor.fetchall()

        # Fallback: se sem resultados com filtros, busca só por família
        if not rows and len(clauses) > 2:
            cursor.execute(
                "SELECT TOP 30 * FROM weg_produtos "
                "WHERE ativo = 1 AND familia = %s "
                "ORDER BY corrente_min, potencia_kvar, potencia_kw, potencia_cv",
                [familia],
            )
            rows = cursor.fetchall()

        conn.close()

        result = []
        for row in rows:
            clean = {}
            for k, v in row.items():
                if v is None:
                    clean[k] = None
                elif hasattr(v, "isoformat"):
                    clean[k] = v.isoformat()
                elif hasattr(v, "__float__") and not isinstance(v, (int, str, bool)):
                    clean[k] = float(v)
                else:
                    clean[k] = v
            result.append(clean)
        return json.dumps(result, ensure_ascii=False)
    except Exception as exc:
        return json.dumps({"erro": str(exc)})


TOOLS = [
    {
        "name": "buscar_produto_weg",
        "description": (
            "Busca produtos WEG no banco de dados Azure SQL. "
            "Use SEMPRE esta ferramenta para encontrar equivalentes WEG antes de responder. "
            "Chame quantas vezes necessário — uma chamada por família de produto identificada. "
            "Retorna lista de produtos com código WAU, SAP, preço, corrente, tensão, kVAr, kW, CV e observações."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "familia": {
                    "type": "string",
                    "description": (
                        "Família WEG: CWM, CWMC, RW, RWM, MPW, MWL, PDW, PDWM, "
                        "CFW100, CFW300, CFW500, CFW11, CFW900, SSW05, SSW07, SSW08, SSW900, "
                        "UCW, UCWT, MCW, BCW, BCWA, PFW03, PFW01, PFWD01, "
                        "CTSW, AHFW, DRW, SPW03, SPW13, PMW01, SWITCH, WCAM"
                    ),
                },
                "corrente_min": {
                    "type": "number",
                    "description": "Corrente mínima desejada em Ampères (faixa inferior)",
                },
                "corrente_max": {
                    "type": "number",
                    "description": "Corrente máxima desejada em Ampères (faixa superior)",
                },
                "tensao_v": {
                    "type": "string",
                    "description": "Tensão exata como string (ex: '220', '380', '440', '480', '24')",
                },
                "potencia_kvar": {
                    "type": "number",
                    "description": "Potência reativa mínima em kVAr (para capacitores)",
                },
                "potencia_kw": {
                    "type": "number",
                    "description": "Potência mínima em kW (para drives, soft-starters)",
                },
                "potencia_cv": {
                    "type": "number",
                    "description": "Potência mínima em CV/HP (para chaves de partida PDW, PDWM)",
                },
                "texto_livre": {
                    "type": "string",
                    "description": "Busca por texto no código, subtipo ou observações do produto",
                },
            },
            "required": ["familia"],
        },
        "cache_control": {"type": "ephemeral"},
    }
]

_DB_INSTRUCTION = (
    "Use SEMPRE a ferramenta `buscar_produto_weg` para consultar o banco de dados WEG. "
    "IMPORTANTE: chame MÚLTIPLAS ferramentas em paralelo na mesma resposta — uma chamada por família identificada, "
    "todas ao mesmo tempo. Nunca faça uma chamada por vez quando houver múltiplas famílias na lista. "
    "Nunca responda sem consultar o banco — os dados não estão no contexto, estão no banco. "
    "Famílias disponíveis: CWM, CWMC, RW, RWM, MPW, MWL, PDW, PDWM, CFW100, CFW300, CFW500, CFW11, CFW900, "
    "SSW05, SSW07, SSW08, SSW900, UCW, UCWT, MCW, BCW, BCWA, PFW03, PFW01, PFWD01, "
    "CTSW, AHFW, DRW, SPW03, SPW13, PMW01, SWITCH, WCAM, CSW, CEW."
)

# ─── System prompt ────────────────────────────────────────────────────────────
_PROMPT_PART1 = """Você é um especialista técnico sênior em produtos elétricos e industriais WEG/WAU, com profundo conhecimento em acionamentos, proteção, automação, CFTV e energia.

Fabricantes concorrentes que você conhece em detalhe: Siemens, ABB, Schneider Electric, Eaton (Moeller), Lovato, Danfoss, Rockwell (Allen-Bradley), Mitsubishi, Yaskawa, WEG antigos.

==========================================================================
## BASE DE DADOS WEG 2026
"""

_PROMPT_PART2 = """
==========================================================================
## REGRAS GERAIS DE CONVERSÃO

1. CORRENTE NOMINAL: a corrente de operação é o critério primário. Se a corrente WEG for menor que a do concorrente para a mesma função, usar o modelo WEG com corrente IMEDIATAMENTE SUPERIOR e anotar: "⚠️ Corrente WEG utilizada: Xa (concorrente: Ya) — confirmar com cliente."
2. CUSTO-BENEFÍCIO: quando houver mais de uma opção WEG tecnicamente válida, SEMPRE indicar a de menor custo/gama que atenda os requisitos. Mencionar em "observacao" se há alternativa premium.
3. ACESSÓRIOS: identificar acessórios solicitados (contatos auxiliares, bobinas, módulos de comunicação, encoders, displays) e indicar os códigos WEG complementares no campo "acessorios_weg".
4. TENSÃO DA BOBINA / ALIMENTAÇÃO: sempre verificar e casar (24 VCC, 24 VCA, 110 V, 220 V, 380 V, 440 V, 480 V).
5. CONCORRENTES — REGRA ABSOLUTA: JAMAIS mencionar, sugerir ou citar produtos de fabricantes concorrentes (Siemens, ABB, Schneider, Eaton, etc.) em qualquer campo da resposta — nem em "observacao", nem em "descricao_weg", nem em "observacoes_gerais". A THP Representações trabalha EXCLUSIVAMENTE com produtos WEG/WAU. Quando não houver equivalente WEG no banco de dados, retornar status="não encontrado" e observacao="Produto não localizado na linha WEG atual — consultar departamento comercial WEG para verificar disponibilidade ou equivalente." NUNCA sugerir que o cliente procure produto de outro fabricante.

==========================================================================
## MATCHING TÉCNICO POR FAMÍLIA DE PRODUTO

### A. DISJUNTORES DE PROTEÇÃO DE MOTOR (DPM / Motor Starter Protectors)
Concorrentes: Siemens 3RV2, 3RV1 | Schneider GV2ME, GV3P | ABB MS116, MS132, MS165 | Eaton PKZM0, PKM0

Parâmetros críticos a extrair e casar:
- Corrente de ajuste (Ir) em Ampères
- Corrente de curto (Icu/Ics) em kA
- Número de polos (3P)
- Versão com ou sem acessórios:
  - Contato auxiliar NA+NF (Side mount ou front mount)
  - Bobina de abertura / shunt trip (bobina de disparo por tensão)
  - Relé de mínima tensão / undervoltage release
  - Bloco diferencial

WEG – linha MPW:
- MPW65-3-D... (até 65 kA) e MPW100-3-D... (até 100 kA)
- Sufixo indica faixa de corrente (ex: D040 = 25–40 A)
- Acessórios WEG: CW1 (contato NA), CW2 (contato NF), BST (bobina shunt), BVM (mínima tensão)
- Indicar sempre código base + códigos de acessórios separados

### B. CONTATORES (Contactors)
Concorrentes: Siemens 3RT2 | Schneider LC1-D, LC1-F | ABB AF-line, A-line | Eaton DILM | Lovato BF/BG

Parâmetros críticos:
- Corrente nominal AC-3 (ex: 9 A, 12 A, 18 A, 25 A, 40 A, 65 A, 80 A, 115 A, 150 A...)
- Tensão da bobina (24 VCC, 24 VCA, 48 V, 110 V, 220 V, 380 V)
- Polos (3P ou 4P)
- Contatos auxiliares integrados: quantos NA e quantos NF
- Se cliente pede contatos adicionais: indicar blocos auxiliares CW1/CW2 separados

WEG – linha CWM:
- CWM09 a CWM150 (corrente AC-3)
- Sufixo bobina: -11 (NA+NF integrado), código completo inclui tensão
- Exemplo: CWM40-11-30V04 = 40A, 1NA+1NF, bobina 220V 50/60Hz

### C. RELÉS DE SOBRECARGA (Overload Relays)
Concorrentes: Siemens 3RU2 | Schneider LRD | ABB TA25DU, TA75DU | Eaton ZB

Parâmetros críticos:
- Faixa de corrente de ajuste
- Reset: manual (M) ou automático (A)
- Classe de disparo: 10, 20 ou 30
- Número de contatos NA/NF

WEG – linha RW:
- RW27D (faixa até 27 A) e RW67D (faixa maior)
- Indicar faixa exata e modo de reset

### D. INVERSORES DE FREQUÊNCIA (Variable Frequency Drives / VFDs)
Concorrentes: Siemens SINAMICS G110/G120/G130/G150 | ABB ACS355/ACS550/ACS880 | Schneider ATV310/ATV312/ATV630/ATV930 | Danfoss FC51/FC102/FC202/FC301/FC302 | Yaskawa V1000/J1000/A1000 | Rockwell PowerFlex 4/40/400/525/755

Parâmetros críticos (TODOS devem ser casados):
- Potência nominal (kW ou HP/CV)
- Tensão de alimentação: 1F 200-240V / 3F 200-240V / 3F 380-480V / 3F 500-600V
- Corrente de saída nominal
- Grau de proteção (IP20, IP21, IP55, IP66)
- Sobrecarga suportada (150% por 60s = uso geral; 110% = bomba/ventilador; 200% = alta performance)

Acessórios/Opções — identificar no código ou descrição do cliente:
- Módulo de comunicação: Profibus DP | Profinet | DeviceNet | EtherNet/IP | CANopen | Modbus TCP (atenção: Modbus RTU já incluso na maioria dos CFW)
- Entradas/saídas digitais adicionais (I/O expansion)
- Encoder / PG card (para controle vetorial com encoder)
- Filtro EMC integrado / externo (categoria C2 ou C3)
- Resistor de frenagem externo (RFD)
- Reator de linha / filtro de saída (bobina dV/dt)
- Display remoto / IHM remota

WEG – linhas CFW:
- CFW500: compacto monofásico/trifásico, uso geral, mais econômico (até ~22 kW)
- CFW700: linha compacta e econômica para bomba/ventilador (até ~250 kW) — PREFERIR para carga quadrática
- CFW11: uso geral robusto, ampla gama de potência (até 2400 kW)
- CFW900 / CFW11: alta performance, vetor com encoder (servo e uso industrial intensivo)
- Seleção de custo-benefício: CFW500 < CFW700 < CFW11 < CFW900. Usar o menor que atender.
- Módulos de comunicação WEG: CFW-11 com slot para módulos MBP (Profibus), CAC (CANopen), ELP (EtherNet/IP), PNT (Profinet), DNet (DeviceNet)

### E. SOFT STARTERS (Chaves de Partida Suave)
Concorrentes: Schneider ATS22/ATS48 | ABB PSR/PSE/PSTB | Siemens 3RW30/3RW40/3RW50 | Eaton DS7

Parâmetros críticos:
- Corrente nominal de saída
- Bypass interno integrado (sim/não — impacta tamanho e custo)
- Tensão de controle
- Módulo de comunicação (se houver)

WEG:
- SSW-07: linha mais econômica, sem bypass
- SSW900: robusta, com bypass opcional, módulos de comunicação disponíveis

### F. RELÉS E TEMPORIZADORES DE PROTEÇÃO
Concorrentes: Schneider RM | ABB CR-P | Siemens 3RP | Phoenix Contact

- Relés temporizadores WEG: linha TW (multifunção, ON-delay, OFF-delay, pulso)
- Relés de falta de fase / desequilíbrio: linha RFW
- Relés de proteção de motor: linha RPM

### G. CHAVES DE PARTIDA DIRETA — SOLUÇÃO COMPLETA EM CAIXA (DOL Starters)
Concorrentes: **Altronic 3PDA** | Siemens 3RA2 | Schneider TeSys starter assemblies | ABB starters in enclosures

ATENÇÃO: São SOLUÇÕES COMPLETAS em caixa metálica (não componentes separados).
Incluem: contator + relé de sobrecarga + botoeira liga/desliga + chave seccionadora.
NÃO sugerir MPW + CWM separados — indicar WEG PDW ou PDWM conforme tensão/fase.

Parâmetros críticos (extrair de código ou descrição):
- Potência do motor (CV/HP)
- Tensão de alimentação (220V monofásica → PDWM | 380V trifásica → PDW)
- Corrente de ajuste do relé de sobrecarga (ex: 12–18A)
- Com ou sem botoeira (sufixo "C/ BOT." ou "B")
- Contato de contorno (CONT): contato auxiliar para sinalização

**LÓGICA DE MATCHING — SIGA ESTA ORDEM:**
1. **Potência (CV) + tensão são critérios primários.** Encontre o WEG com o mesmo CV e mesma tensão.
2. **Faixa do relé pode diferir** entre fabricantes — é normal. Selecione o WEG de mesmo CV e anote a diferença na observação: "⚠ Faixa WEG: X-Y A — confirmar ajuste com cliente."
3. **Só escale para CV maior** se o WEG de mesma potência genuinamente não cobre a corrente nominal do motor (faixa máxima menor que a corrente nominal informada). Nesse caso marque como alerta e explique.

**TABELA PDW — TRIFÁSICA 380V (V40) — CÓDIGOS SAP CONFIRMADOS:**
| Referência WEG   | Potência (CV) | Faixa (A)    | Código SAP |
|------------------|--------------|--------------|------------|
| PDW02-0,16V40    | 0,16         | 0,4 – 0,63  | 10072580   |
| PDW02-0,25V40    | 0,25         | 0,56 – 0,8  | 10186081   |
| PDW02-0,33V40    | 0,33         | 0,8 – 1,2   | 10186082   |
| PDW02-0,75V40    | 0,5 / 0,75   | 1,2 – 1,8   | 10045784   |
| PDW02-1,5V40     | 1 / 1,5      | 1,8 – 2,8   | 10118384   |
| PDW02-2V40       | 2            | 2,8 – 4     | 10045787   |
| PDW02-3V40       | 3            | 4 – 6,3     | 10045788   |
| PDW02-4V40       | 4            | 5,6 – 8     | 10045789   |
| PDW04-5V40       | 5            | 7 – 10      | 10045790   |
| PDW04-6V40       | 6            | 8 – 12,5    | 10045791   |
| PDW04-7,5V40     | 7,5          | 10 – 15     | 10045792   |
| PDW04-10V40      | 10           | 11 – 17     | 10045793   |
| PDW04-12,5V40    | 12,5         | 15 – 23     | 10045794   |
| PDW04-15V40      | 15           | 22 – 32     | 10046425   |

**TABELA PDWM — MONOFÁSICA 220V (V25) — CÓDIGOS SAP CONFIRMADOS:**
| Referência WEG         | Potência (CV)  | Faixa (A)   | Código SAP |
|------------------------|---------------|-------------|------------|
| PDWM02-0,16/0,125V25   | 0,125 / 0,16  | 1,2 – 1,8  | 10070900   |
| PDWM02-0,33V25         | 0,25 / 0,33   | 2,8 – 4    | 10046171   |
| PDWM02-0,5/0,75V25     | 0,5 / 0,75    | 4 – 6,3    | 10046170   |
| PDWM02-1V25            | 0,75 / 1      | 5,6 – 8    | 10118357   |
| PDWM04-1,5AV25         | 1,5           | 7 – 10     | 10907057   |
| PDWM04-2A/1,5NV25      | 2             | 8 – 12,5   | 10045728   |
| PDWM04-3/2V25          | 3             | 11 – 17    | 10045729   |
| PDWM04-4AV25           | 4             | 15 – 23    | 10045739   |
| PDWM04-5AV25           | 5             | 22 – 32    | 10046444   |
| PDWM05-2A/1,5NV25      | 2             | 8 – 12,5   | 13339233   |
| PDWM05-3/2V25          | 3             | 11 – 17    | 13339229   |
| PDWM05-4AV25           | 4             | 15 – 23    | 13339270   |
| PDWM05-5AV25           | 5             | 22 – 32    | 13339236   |
| PDWM05-7,5AV25         | 7,5           | 32 – 40    | 13336722   |
| PDWM06-7,5AV25         | 7,5           | 32 – 40    | 10045740   |
| PDWM08-10AV25          | 10            | 32 – 50    | 10045766   |
| PDWM08-12,5AV25        | 12,5          | 40 – 57    | 10045741   |
| PDWM08-15AV25          | 15            | 57 – 70    | 10046182   |

Notas: V25=220V CA | V40=380V CA | V49=440V CA | VC8=127V CA
Se a faixa do relé WEG for menor que a do concorrente para o mesmo CV: selecione o modelo com faixa imediatamente superior e anote "⚠ Faixa WEG: X-Y A — confirmar ajuste com cliente".
Estes códigos SAP são confirmados — status="encontrado" quando o CV e tensão baterem exatamente.

### H. PROTETORES DE SURTO (SPDs — Surge Protection Devices)
- Usar a ferramenta buscar_produto_weg com família SPW03 ou SPW13 como prioridade
- Para itens não listados: casar por tensão (Un), corrente máxima (Imax/kA), classe (I, II, III, I/II), tecnologia (varistor MOV, ECG)
- WEG: SPW03 (CA classe II), SPW13 (CC/fotovoltaica)

### I. EQUIPAMENTOS CFTV / REDE
- Câmeras: casar por resolução (MP), distância IR (m), tipo (bullet/dome/PTZ), IP65/66, protocolo ONVIF
- Switches: casar por número de portas PoE e não-PoE, potência total PoE, velocidade (10/100/1000)
- Usar base WAU 14/2026 e 15/2026

### J. CAPACITORES E CORREÇÃO DO FATOR DE POTÊNCIA
**ATENÇÃO: WEG FABRICA linha completa de capacitores. NUNCA retornar "não encontrado" sem primeiro consultar o banco com buscar_produto_weg.**

Famílias WEG e quando usar cada uma:
- **UCW** (Unidade Capacitiva Monofásica): capacitor individual monofásico. Identificar: kVAr + tensão (220/380/440/480/535V). Código: UCW{kVAr}V{tensao} + sufixo tamanho.
- **UCWT HD** (Unidade Capacitiva Trifásica): capacitor individual trifásico padrão. Usar para maioria dos itens de capacitor trifásico. Código: UCWT{kVAr}V{tensao} L/N/Q/S/U HD.
- **UCWT UHD** (Ultra Heavy Duty): versão reforçada do UCWT, para uso com chaves tiristorizadas ou harmônicos elevados.
- **MCW** (Módulo Capacitivo Trifásico): banco pré-montado com UCWs integradas. Mais compacto e prático que UCWT isolado. Usar quando cliente pede "banco modular" ou "módulo capacitivo".
- **BCW** (Banco de Capacitores Trifásico em caixa): banco fechado com caixa metálica. Usar para instalações simples.
- **BCWP-D** (Banco com Disjuntor): BCW com disjuntor integrado.
- **BCWA** (Banco Automático): banco com controlador automático integrado (PFW03). Usar quando cliente pede "banco automático de capacitores" completo.
- **PFW03/PFW01**: controladores automáticos de fator de potência separados. PFW03=50/60Hz universal, PFW01=60Hz ou 50Hz específico.
- **PFWD01**: controlador dinâmico (para cargas variáveis rápidas, thyristor output).
- **CTSW**: chave tiristorizada para manobra de capacitores sem contator. Substitui Epcos/TDK PhaseCap switch, Nokian NKKTS, FRAKO KM.
- **AHFW**: filtro ativo para correção de FP e harmônicos. Substitui ABB PQF, Schneider AccuSine, Siemens SiCap.
- **DRW**: reator de dessintonia, sempre em conjunto com UCWT. Necessário quando há harmônicos na rede (redes com drives/inversores).

Parâmetros críticos para matching de capacitores:
1. **kVAr** (potência reativa): critério primário. Casar exatamente ou imediatamente acima.
2. **Tensão nominal** (V): 220V=V25 | 380V=V40 | 440V=V49 | 480V=V53 | 535V=V57 | 660V=V63 | 600V=V103
3. **Fases**: monofásico → UCW; trifásico → UCWT ou MCW
4. **Tipo de produto**: capacitor avulso (UCWT), banco (BCW/MCW), automático (BCWA), controlador (PFW), chave (CTSW), filtro ativo (AHFW)

Concorrentes mais comuns de capacitores:
- **ABB**: CLMD (trifásico), CLMB (monofásico), PQF (filtro ativo)
- **Schneider/Capacitor Industries**: Varplus Can, Vlarpluselec, AccuSine (filtro ativo)
- **Epcos / TDK**: MKV, B32 (monofásico), PhaseCap HD (trifásico), PhaseCap Speed (com chave tiristorizada)
- **Nokian**: NKK, NKKT (trifásico), NKKTS (chave tiristorizada)
- **FRAKO**: EM, KM (chave), LM, PM
- **Ducati / Icar / Comar**: CPT, MKP, capacitores italianos
- **Controladores concorrentes**: ABB RVT, Schneider Varplus Box, Lovato DCRG, Selec, Ducati MC

Regra de matching por item de texto livre (ex: "Capacitor 16,24 kvar 380 volts"):
→ 16,24 kVAr trifásico 380V → buscar com buscar_produto_weg(familia="UCWT", potencia_kvar=16.24, tensao_v="380")
Regra: se kVAr exato não existe, usar UCWT imediatamente acima e registrar no campo "observacao": "⚠ kVAr WEG imediatamente superior — confirmar com cliente".

==========================================================================
### K. SINALIZAÇÃO E COMANDO (Pilot Lights / Sinaleiros / Botoeiras / Comutadores)
Concorrentes identificados por código: Siemens 3SB3 | Schneider XB7, ZB4, ZB5 | Eaton M22, RMQ | ABB CP, CP-S | Lovato LP

WEG — famílias CSW e CEW (Sinalização e Comando Ø22mm):
- Sinaleiros/lâmpadas piloto LED: buscar com buscar_produto_weg(familia="CSW", texto_livre="sinaleiro LED <cor> <tensão>V")
- Botoeiras pulsadoras: buscar com buscar_produto_weg(familia="CSW", texto_livre="botao pulsador <cor>")
- Botoeiras iluminadas: buscar com buscar_produto_weg(familia="CSW", texto_livre="botao iluminado <cor> <tensão>V")
- Botões de emergência: buscar com buscar_produto_weg(familia="CSW", texto_livre="botao emergencia")
- Comutadores/seletores: buscar com buscar_produto_weg(familia="CSW", texto_livre="comutador seletor")
- CEW (versão Modular): buscar com buscar_produto_weg(familia="CEW", texto_livre="<tipo> <cor>")

Parâmetros críticos a extrair e casar:
- Cor: vermelho / verde / amarelo / branco / azul / laranja
- Tensão: 24V CC | 24V CA | 110V CA | 220V CA | 127V CA
- Tipo: LED (preferencial) ou lâmpada incandescente/néon
- Montagem: 22mm (padrão industrial) ou 30mm
- Contatos auxiliares necessários: NA e/ou NF

ATENÇÃO: Se não encontrar no banco, retornar status="não encontrado" com observacao="Consultar departamento comercial WEG — linha CSW/CEW de sinalização e comando 22mm disponível em diversas cores e tensões." NUNCA sugerir produto de concorrente.

==========================================================================
## IDENTIFICAÇÃO DE FABRICANTE POR CÓDIGO

- Siemens: 3RT, 3RV, 3RU, 3RP, 3SB, SINAMICS (G120, S120), 6SL, 6ES, 6GK, 3RA2 (combinações partida direta)
- ABB: AF, A (A9..A300), MS, TA, ACS, ACH, ACS355/550/880, S200, SH200
- Schneider: LC1, LC2, LRD, GV2, GV3, ATV, ATS, NSX, CVS, RH, RM, LX, XB
- Eaton (Moeller): DILM, DILK, ZB, PKZM, PKM, NZM, DS7
- Danfoss: FC51, FC102, FC202, FC301, FC302, VLT
- Rockwell / Allen-Bradley: 100-C, 140M, 509, 520, 525, 755, 22B, 25B
- Lovato: BF, BG, BX, RGK
- Altronic: 3PDA (chaves de partida direta em caixa metálica — prefixo 3PDA + números de potência/tensão + sufixo B=com botoeira)
- WEG antigo: WSW, WEG-CFW (linha antiga)
- Capacitores — Epcos/TDK: B3232, MKV, PhaseCap, B32K (capacitores de potência)
- Capacitores — Nokian: NKK, NKKT, NKKF, NKKTS (chave tiristorizada)
- Capacitores — ABB: CLMD, CLMB, RVT (controlador)
- Capacitores — FRAKO: EM, KM, LM, PM
- Capacitores — Icar / Ducati / Comar / Elco: CPT, MKP (italianos)
- Capacitores — Schneider: Varplus Can, Vlarpluselec
- Controladores de FP — ABB: RVT | Schneider: Varplus Box | Lovato: DCRG | Selec: PFCR

==========================================================================
## FORMATO DE RESPOSTA — RETORNE APENAS JSON VÁLIDO, SEM TEXTO EXTRA:

{
  "items": [
    {
      "seq": 1,
      "codigo_cliente": "código exatamente como recebido",
      "quantidade": 1,
      "fabricante": "fabricante identificado",
      "descricao_cliente": "descrição/nome do produto do cliente",
      "tipo_produto": "categoria: Contator / DPM / Drive / Soft Starter / SPD / Câmera / etc.",
      "especificacoes": "corrente, tensão alimentação, tensão bobina, polos, acessórios, grau proteção, comunicação, etc. — COMPLETO",
      "codigo_weg": "código WAU 8 dígitos (item principal) ou vazio",
      "referencia_weg": "referência WEG (ex: CWM40-11-30V04) ou vazio",
      "descricao_weg": "descrição completa do produto WEG equivalente",
      "acessorios_weg": "códigos/referências WEG de acessórios separados necessários (ex: CW1, BVM, módulo ELP) ou vazio",
      "preco_lista": "valor em R$ somente se retornado pelo banco de dados; senão vazio",
      "observacao": "⚠️ avisos: corrente superior utilizada, acessório não disponível na WEG, confirmar tensão de bobina, alternativa premium disponível, etc.",
      "status": "encontrado"
    }
  ],
  "resumo": {
    "total_itens": 0,
    "encontrados": 0,
    "parciais": 0,
    "nao_encontrados": 0,
    "observacoes_gerais": "observações gerais sobre a lista"
  }
}

Valores possíveis para "status": "encontrado" | "parcial" | "não encontrado"
"""

SYSTEM_PROMPT = _PROMPT_PART1 + _DB_INSTRUCTION + _PROMPT_PART2

# System prompt em formato de cache — economiza ~70% nos tokens de entrada
# Cache dura 5 min no servidor Anthropic, compartilhado entre requisições do mesmo API key
_CACHED_SYSTEM = [
    {
        "type": "text",
        "text": SYSTEM_PROMPT,
        "cache_control": {"type": "ephemeral"},
    }
]

# ─── Helpers ─────────────────────────────────────────────────────────────────
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic", ".heif", ".bmp", ".tiff", ".tif"}
EXCEL_EXTS = {".xlsx", ".xls"}
CSV_EXTS   = {".csv"}

MEDIA_TYPES = {
    ".png": "image/png", ".gif": "image/gif", ".webp": "image/webp",
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".bmp": "image/jpeg",
    ".tiff": "image/jpeg", ".tif": "image/jpeg",
    ".heic": "image/jpeg", ".heif": "image/jpeg",
}


def parse_json_response(text: str) -> dict:
    text = re.sub(r"```json\s*", "", text)
    text = re.sub(r"```\s*", "", text)
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass
    return {
        "items": [],
        "resumo": {
            "total_itens": 0, "encontrados": 0, "parciais": 0,
            "nao_encontrados": 0,
            "observacoes_gerais": "Não foi possível interpretar a resposta. Tente novamente.",
        },
    }


def _execute_tool(tool_name: str, tool_input: dict) -> str:
    if tool_name == "buscar_produto_weg":
        return buscar_produto_weg(**tool_input)
    return json.dumps({"erro": "Ferramenta desconhecida: " + tool_name})


def call_claude_image(image_bytes: bytes, ext: str) -> dict:
    image_b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    media_type = MEDIA_TYPES.get(ext, "image/jpeg")
    messages = [
        {
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {"type": "base64", "media_type": media_type, "data": image_b64},
                },
                {
                    "type": "text",
                    "text": (
                        "Analise esta imagem com lista de materiais elétricos/industriais de concorrentes. "
                        "Extraia TODOS os códigos e quantidades — mesmo que a escrita seja difícil, faça o melhor. "
                        "Para cada item, identifique o fabricante, extraia TODAS as especificações técnicas visíveis "
                        "(corrente, tensão, acessórios, comunicação, etc.) e converta para equivalente WEG usando "
                        "a ferramenta buscar_produto_weg conforme as regras do sistema. "
                        "RETORNE APENAS o JSON, sem texto adicional."
                    ),
                },
            ],
        }
    ]
    for _ in range(8):
        resp = client.messages.create(
            model=VISION_MODEL,
            max_tokens=8192,
            system=_CACHED_SYSTEM,
            tools=TOOLS,
            messages=messages,
        )
        if resp.stop_reason == "end_turn":
            for block in resp.content:
                if hasattr(block, "text"):
                    return parse_json_response(block.text)
            break
        if resp.stop_reason == "tool_use":
            messages.append({"role": "assistant", "content": resp.content})
            tool_results = []
            for block in resp.content:
                if block.type == "tool_use":
                    result = _execute_tool(block.name, block.input)
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": result,
                    })
            messages.append({"role": "user", "content": tool_results})
            continue
        break
    return parse_json_response("")


def call_claude_text(codes_text: str, source_hint: str = "planilha") -> dict:
    messages = [
        {
            "role": "user",
            "content": (
                "Lista de materiais do cliente (origem: " + source_hint + "):\n\n"
                "```\n" + codes_text + "\n```\n\n"
                "Para cada item: identifique o fabricante pelo código, extraia TODAS as especificações técnicas "
                "(corrente, tensão, acessórios, tipo de contato, protocolo de comunicação, grau de proteção, etc.) "
                "e converta para o equivalente WEG de melhor custo-benefício usando a ferramenta buscar_produto_weg. "
                "RETORNE APENAS o JSON, sem texto adicional."
            ),
        }
    ]
    for _ in range(8):
        resp = client.messages.create(
            model=TEXT_MODEL,
            max_tokens=8192,
            system=_CACHED_SYSTEM,
            tools=TOOLS,
            messages=messages,
        )
        if resp.stop_reason == "end_turn":
            for block in resp.content:
                if hasattr(block, "text"):
                    return parse_json_response(block.text)
            break
        if resp.stop_reason == "tool_use":
            messages.append({"role": "assistant", "content": resp.content})
            tool_results = []
            for block in resp.content:
                if block.type == "tool_use":
                    result = _execute_tool(block.name, block.input)
                    tool_results.append({
                        "type": "tool_result",
                        "tool_use_id": block.id,
                        "content": result,
                    })
            messages.append({"role": "user", "content": tool_results})
            continue
        break
    return parse_json_response("")


def process_image(content: bytes, ext: str) -> dict:
    if ext in {".heic", ".heif", ".tiff", ".tif", ".bmp"}:
        try:
            from PIL import Image as PILImage
            buf = io.BytesIO(content)
            img = PILImage.open(buf)
            out = io.BytesIO()
            img.convert("RGB").save(out, format="JPEG", quality=90)
            out.seek(0)
            content = out.read()
            ext = ".jpg"
        except Exception:
            pass
    return call_claude_image(content, ext)


def _extrair_codigo_weg_descricao(texto: str):
    """Extrai código SAP WEG (7-10 dígitos) de parênteses na descrição."""
    if not texto:
        return None
    matches = re.findall(r'\((\d{7,10})\)', str(texto))
    return matches[-1] if matches else None


def _detectar_tabela_excel(content: bytes):
    """Detecta cabeçalho real da tabela e extrai itens com códigos WEG embutidos."""
    try:
        import openpyxl as _opxl
        wb = _opxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        KEYWORDS = {'cod', 'código', 'codigo', 'item', 'descriç', 'descrip', 'quant', 'unid', 'ref', 'material'}
        header_row_idx = None
        for i, row in enumerate(all_rows):
            row_text = ' '.join(str(v).lower() for v in row if v is not None).strip()
            if row_text and sum(1 for kw in KEYWORDS if kw in row_text) >= 2:
                header_row_idx = i
                break

        if header_row_idx is None:
            return None

        headers = [str(v).strip() if v else "Col" + str(j) for j, v in enumerate(all_rows[header_row_idx])]
        rows_data = []
        for row in all_rows[header_row_idx + 1:]:
            if any(v is not None and str(v).strip() for v in row):
                rows_data.append({headers[j]: (str(v).strip() if v is not None else '') for j, v in enumerate(row)})

        if not rows_data:
            return None

        desc_col = next((k for k in headers if 'desc' in k.lower()), None)
        lines = []
        seq = 0
        for row in rows_data:
            if not any(v.strip() for v in row.values()):
                continue
            seq += 1
            desc_text = row.get(desc_col, '') if desc_col else ''
            weg_code = _extrair_codigo_weg_descricao(desc_text)
            parts = ["Item " + str(seq) + ":"]
            for k, v in row.items():
                if v.strip():
                    parts.append(k + "=" + v)
            if weg_code:
                parts.append("[CÓDIGO SAP WEG NA DESCRIÇÃO: " + weg_code + "]")
                parts.append("[PRODUTO JÁ É WEG — preencher codigo_weg com este valor; extrair referencia_weg do modelo na descrição; status=encontrado]")
            lines.append(' | '.join(parts))

        return '\n'.join(lines) if lines else None
    except Exception:
        return None


def process_excel(content: bytes) -> dict:
    # Detecção inteligente: lida com planilhas de pedido/cotação formatadas
    enriched = _detectar_tabela_excel(content)
    if enriched:
        return call_claude_text(enriched, "Excel (tabela auto-detectada com códigos WEG extraídos)")

    # Fallback: pandas direto
    try:
        df = pd.read_excel(io.BytesIO(content), engine="openpyxl")
    except Exception:
        df = pd.read_excel(io.BytesIO(content))
    return call_claude_text(df.to_string(index=False), "Excel")


def process_csv(content: bytes) -> dict:
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            df = pd.read_csv(io.BytesIO(content), encoding=enc)
            break
        except Exception:
            continue
    else:
        raise ValueError("Não foi possível decodificar o CSV.")
    return call_claude_text(df.to_string(index=False), "CSV")


# ─── Excel generator ──────────────────────────────────────────────────────────
WEG_BLUE   = "00205B"
WEG_GREEN  = "00853D"
LIGHT_BLUE = "E8F0FE"


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def generate_excel(items: list, resumo: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conversão WEG"

    # Title
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "CONVERSÃO DE CÓDIGOS WEG  –  THP Representações"
    c.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    c.fill = PatternFill(start_color=WEG_BLUE, end_color=WEG_BLUE, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # Headers
    COLS = [
        ("#",                   4),
        ("Código Cliente",     18),
        ("Qtd",                 5),
        ("Fabricante",         14),
        ("Tipo",               14),
        ("Especificações",     30),
        ("Código WEG",         16),
        ("Referência WEG",     22),
        ("Descrição WEG",      34),
        ("Acessórios WEG",     22),
        ("Preço Lista (R$)",   14),
        ("Observação",         40),
        ("Status",             16),
    ]
    for col_i, (label, width) in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=col_i, value=label)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = PatternFill(start_color=WEG_BLUE, end_color=WEG_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[2].height = 28

    fills = {
        "encontrado":     PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "parcial":        PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "não encontrado": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    }
    alt_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    status_meta = {
        "encontrado":     ("155724", "✅ Encontrado"),
        "parcial":        ("856404", "⚠️ Parcial"),
        "não encontrado": ("721C24", "❌ Não encontrado"),
    }

    for r_i, item in enumerate(items, 3):
        status = item.get("status", "não encontrado").lower().strip()
        st_color, st_label = status_meta.get(status, ("721C24", "❌ Não encontrado"))

        row_vals = [
            item.get("seq", r_i - 2),
            item.get("codigo_cliente", ""),
            item.get("quantidade", 1),
            item.get("fabricante", ""),
            item.get("tipo_produto", ""),
            item.get("especificacoes", ""),
            item.get("codigo_weg", ""),
            item.get("referencia_weg", ""),
            item.get("descricao_weg", ""),
            item.get("acessorios_weg", ""),
            item.get("preco_lista", ""),
            item.get("observacao", ""),
            st_label,
        ]

        for col_i, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r_i, column=col_i, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border()

            if col_i == 13:   # Status
                cell.fill = fills.get(status, fills["não encontrado"])
                cell.font = Font(name="Calibri", size=10, bold=True, color=st_color)
            elif col_i == 7:  # Código WEG
                cell.font = Font(name="Calibri", size=10, bold=True, color=WEG_BLUE)
            elif col_i == 12 and val:  # Observação
                cell.font = Font(name="Calibri", size=10, italic=True, color="7B5E00")
            elif r_i % 2 == 0:
                cell.fill = alt_fill

        ws.row_dimensions[r_i].height = 38

    # Summary
    s_row = len(items) + 4
    ws.merge_cells("A" + str(s_row) + ":M" + str(s_row))
    sc = ws["A" + str(s_row)]
    sc.value = (
        "Total: " + str(resumo.get('total_itens', len(items))) + "  |  "
        "Encontrados: " + str(resumo.get('encontrados', 0)) + "  |  "
        "Parciais: " + str(resumo.get('parciais', 0)) + "  |  "
        "Não encontrados: " + str(resumo.get('nao_encontrados', 0))
    )
    sc.font = Font(bold=True, name="Calibri", size=10, color=WEG_BLUE)
    sc.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    sc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[s_row].height = 22

    obs_geral = resumo.get("observacoes_gerais", "")
    if obs_geral:
        o_row = s_row + 1
        ws.merge_cells("A" + str(o_row) + ":M" + str(o_row))
        oc = ws["A" + str(o_row)]
        oc.value = "Obs.: " + obs_geral
        oc.font = Font(italic=True, name="Calibri", size=10, color="856404")
        oc.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        oc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[o_row].height = 22

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A2:M" + str(len(items) + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Routes ───────────────────────────────────────────────────────────────────
@app.get("/")
async def root():
    return FileResponse("static/index.html")

@app.get("/health")
async def health():
    try:
        conn = get_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM weg_produtos WHERE ativo = 1")
        total = cursor.fetchone()[0]
        conn.close()
        return {"status": "ok", "db": "connected", "produtos": total}
    except Exception as e:
        return {"status": "ok", "db": "error", "detail": str(e)}

@app.post("/api/convert")
async def convert(file: UploadFile = File(...)):
    content = await file.read()
    name = file.filename or ""
    ext  = Path(name).suffix.lower()

    try:
        if ext in IMAGE_EXTS:
            result = process_image(content, ext)
        elif ext in EXCEL_EXTS:
            result = process_excel(content)
        elif ext in CSV_EXTS:
            result = process_csv(content)
        else:
            raise HTTPException(
                status_code=400,
                detail="Formato '" + (ext or name) + "' não suportado. Use imagem (JPG, PNG, HEIC) ou planilha (XLSX, CSV).",
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Erro ao processar: " + str(exc)) from exc

    return JSONResponse(content=result)

class TextInput(BaseModel):
    texto: str

@app.post("/api/convert-text")
async def convert_text_endpoint(body: TextInput):
    """Converte lista de produtos a partir de texto livre (mensagem WhatsApp, e-mail, etc.)."""
    if not body.texto.strip():
        raise HTTPException(status_code=400, detail="Texto vazio. Cole a mensagem e tente novamente.")
    try:
        result = call_claude_text(body.texto.strip(), "mensagem WhatsApp / texto livre")
    except Exception as exc:
        raise HTTPException(status_code=500, detail="Erro ao processar: " + str(exc)) from exc
    return JSONResponse(content=result)

@app.post("/api/export")
async def export(request: Request):
    data   = await request.json()
    items  = data.get("items", [])
    resumo = data.get("resumo", {})
    buf = generate_excel(items, resumo)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=conversao_weg.xlsx"},
    )
